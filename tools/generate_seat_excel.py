import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

try:
    # 尝试使用openpyxl读取
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    # 尝试使用xlrd读取旧格式xls文件
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
    print("WARNING: xlrd not installed. Old .xls files may not be supported.")
    print("To support .xls files, run: pip install xlrd==1.2.0")


# ========== 颜色定义（和截图完全匹配） ==========
# 深蓝背景（标题/表头）：Excel标准主题蓝
BLUE_BG = PatternFill('solid', fgColor='0070C0')
# 空座浅灰背景
GRAY_BG = PatternFill('solid', fgColor='F2F2F2')

# 白色字体（标题栏用）
white_font = Font(size=11, bold=True, color='FFFFFF')
# 常规黑色字体
normal_font = Font(size=11, color='000000')
# 排数加粗字体
row_bold_font = Font(size=11, bold=True, color='000000')
# 空座灰色字体
gray_font = Font(size=11, color='808080')

# 居中对齐
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 全表细边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 文本格式
text_format = '@'


def set_empty_seat_style(cell):
    """设置空座样式：浅灰背景 + 灰色字体"""
    cell.fill = GRAY_BG
    cell.font = gray_font


def parse_seat(seat_str: str):
    """解析座号，拆分排号和座位字母，例：12A -> (12, 'A')"""
    if not seat_str:
        return None
    match = re.match(r'^(\d+)([A-Fa-f])$', str(seat_str).strip())
    if match:
        return int(match.group(1)), match.group(2).upper()
    return None


def format_date(date_str: str):
    """将20260608格式转为2026/06/08"""
    s = str(date_str).strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}/{s[4:6]}/{s[6:]}"
    return s


def load_passenger_data(input_path: str):
    """读取乘客Excel，支持xlsx和xls格式"""
    passengers = []
    headers = []
    trip_info = {}

    # 检查文件扩展名
    if input_path.lower().endswith('.xls') and not input_path.lower().endswith('.xlsx'):
        # 旧格式xls文件
        if not HAS_XLRD:
            raise ValueError("This is an old .xls file. Please install xlrd: pip install xlrd==1.2.0\nOr convert the file to .xlsx format first.")

        print("INFO: Reading old .xls format file...")
        wb = xlrd.open_workbook(input_path)
        ws = wb.sheet_by_index(0)

        # 读取表头
        headers = [str(ws.cell_value(0, col)).strip() for col in range(ws.ncols)]

        # 读取数据
        for row_idx in range(1, ws.nrows):
            row_data = {}
            for col_idx in range(ws.ncols):
                row_data[headers[col_idx]] = ws.cell_value(row_idx, col_idx)

            name = str(row_data.get('姓名', '')).strip()
            carriage = str(row_data.get('车厢', '')).strip()
            seat = str(row_data.get('座号', '')).strip()

            if not name or not carriage or not seat or name == 'None':
                continue

            seat_info = parse_seat(seat)
            if not seat_info:
                continue

            row_data['_排号'] = seat_info[0]
            row_data['_座位字母'] = seat_info[1]
            passengers.append(row_data)

            if not trip_info:
                trip_info = {
                    '日期': format_date(row_data.get('日期', '')),
                    '车次': str(row_data.get('车次', '')).strip(),
                    '发站': str(row_data.get('发站', '')).strip(),
                    '到站': str(row_data.get('到站', '')).strip()
                }
    else:
        # 新格式xlsx文件
        print("INFO: Reading .xlsx format file...")
        wb = load_workbook(input_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return [], [], {}

        headers = [str(h).strip() if h else '' for h in rows[0]]

        for row in rows[1:]:
            row_data = dict(zip(headers, row))
            name = str(row_data.get('姓名', '')).strip()
            carriage = str(row_data.get('车厢', '')).strip()
            seat = str(row_data.get('座号', '')).strip()

            if not name or not carriage or not seat or name == 'None':
                continue

            seat_info = parse_seat(seat)
            if not seat_info:
                continue

            row_data['_排号'] = seat_info[0]
            row_data['_座位字母'] = seat_info[1]
            passengers.append(row_data)

            if not trip_info:
                trip_info = {
                    '日期': format_date(row_data.get('日期', '')),
                    '车次': str(row_data.get('车次', '')).strip(),
                    '发站': str(row_data.get('发站', '')).strip(),
                    '到站': str(row_data.get('到站', '')).strip()
                }

    return passengers, headers, trip_info


def build_seat_map(passengers: list):
    """按 车厢->排号->座位字母 层级组织数据"""
    carriage_map = {}
    for p in passengers:
        carriage = p['车厢']
        row_num = p['_排号']
        seat_char = p['_座位字母']
        carriage_map.setdefault(carriage, {}).setdefault(row_num, {})[seat_char] = p['姓名']
    return carriage_map


def generate_seat_excel(input_path: str, output_path: str):
    """生成带完整样式的座位表Excel"""
    print(f"INPUT: {input_path}")
    print(f"OUTPUT: {output_path}")
    print()

    # 1. 读取数据
    print("Step 1: Loading passenger data...")
    passengers, headers, trip_info = load_passenger_data(input_path)

    if not passengers:
        raise ValueError("No valid passenger data found in the Excel file")

    print(f"  Found {len(passengers)} passengers")

    carriage_map = build_seat_map(passengers)
    print(f"  Found {len(carriage_map)} carriages")

    # 2. 使用全局样式常量
    print("\nStep 2: Creating Excel workbook...")

    wb = Workbook()

    # ========== 工作表1：乘客车票信息 ==========
    print("  Creating passenger info sheet...")
    ws1 = wb.active
    ws1.title = "Passenger Info"

    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = white_font
        cell.fill = BLUE_BG
        cell.alignment = center_align
        cell.border = thin_border

    text_cols = ['证件号码', '订单号', '窗口号']
    for row_idx, p in enumerate(passengers, 2):
        for col_idx, header in enumerate(headers, 1):
            value = p.get(header, '')
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = normal_font
            if header in text_cols:
                cell.number_format = text_format

    for col_idx in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 16
    ws1.freeze_panes = 'A2'

    # ========== 工作表2：列车座位表（7列标准布局） ==========
    print("  Creating seat table sheet...")
    ws2 = wb.create_sheet("Seat Table")

    # 精准列宽（和截图比例一致）
    col_width_map = {
        'A': 12,   # 排数列
        'B': 28,   # A座
        'C': 28,   # B座
        'D': 28,   # C座
        'E': 8,    # 过道列
        'F': 28,   # D座
        'G': 28    # F座
    }
    for col_letter, width in col_width_map.items():
        ws2.column_dimensions[col_letter].width = width

    current_row = 1
    left_seats = ['A', 'B', 'C']
    right_seats = ['D', 'F']

    # 生成每个车厢
    print("  Generating seat layout...")
    for carriage in sorted(carriage_map.keys()):
        row_map = carriage_map[carriage]
        min_row = min(row_map.keys())
        max_row = max(row_map.keys())
        row_count = max_row - min_row + 1

        # ========== 1. 车厢标题行 ==========
        # 组合标题：日期 行程 车次 车厢号 座位表
        date_str = trip_info.get('日期', '')
        route_str = f"{trip_info.get('发站', '')}-{trip_info.get('到站', '')}"
        train_no = trip_info.get('车次', '')

        title_parts = [date_str, route_str, train_no, f"{carriage}号车厢 座位表"]
        carriage_name = ' '.join([p for p in title_parts if p])

        ws2.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        title_cell = ws2.cell(row=current_row, column=1)
        title_cell.value = carriage_name
        title_cell.font = white_font
        title_cell.fill = BLUE_BG
        title_cell.alignment = center_align
        title_cell.border = thin_border
        ws2.row_dimensions[current_row].height = 28

        # ========== 2. 表头行 ==========
        header_row = current_row + 1

        # 三人座：合并B-D列
        ws2.merge_cells(start_row=header_row, start_column=2, end_row=header_row, end_column=4)
        cell_three = ws2.cell(row=header_row, column=2)
        cell_three.value = "三人座（靠窗→过道）"
        cell_three.font = white_font
        cell_three.fill = BLUE_BG
        cell_three.alignment = center_align
        cell_three.border = thin_border

        # 过道列：E列
        cell_aisle = ws2.cell(row=header_row, column=5)
        cell_aisle.value = "过道"
        cell_aisle.font = white_font
        cell_aisle.fill = BLUE_BG
        cell_aisle.alignment = center_align
        cell_aisle.border = thin_border

        # 二人座：合并F-G列
        ws2.merge_cells(start_row=header_row, start_column=6, end_row=header_row, end_column=7)
        cell_two = ws2.cell(row=header_row, column=6)
        cell_two.value = "二人座（过道→靠窗）"
        cell_two.font = white_font
        cell_two.fill = BLUE_BG
        cell_two.alignment = center_align
        cell_two.border = thin_border

        # 排数列表头（A列）
        cell_row_header = ws2.cell(row=header_row, column=1)
        cell_row_header.fill = BLUE_BG
        cell_row_header.border = thin_border

        ws2.row_dimensions[header_row].height = 24

        # ========== 3. 数据行 ==========
        data_start_row = header_row + 1
        for seat_row in range(min_row, max_row + 1):
            seats = row_map.get(seat_row, {})
            data_row = data_start_row + (seat_row - min_row)

            # A列：排数（加粗）
            cell = ws2.cell(row=data_row, column=1, value=f"{seat_row}排")
            cell.font = row_bold_font
            cell.alignment = center_align
            cell.border = thin_border

            # B-D列：左侧三人座 A/B/C
            for idx, seat_char in enumerate(left_seats):
                col = 2 + idx
                name = seats.get(seat_char, '')
                cell_val = f"{seat_row}{seat_char} {name}" if name else f"{seat_row}{seat_char}"
                cell = ws2.cell(row=data_row, column=col, value=cell_val)
                cell.alignment = center_align
                cell.border = thin_border

                # 空座样式
                if not name:
                    set_empty_seat_style(cell)
                else:
                    cell.font = normal_font

            # E列：过道（破折号）
            cell = ws2.cell(row=data_row, column=5, value="——")
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = normal_font

            # F-G列：右侧二人座 D/F
            for idx, seat_char in enumerate(right_seats):
                col = 6 + idx
                name = seats.get(seat_char, '')
                cell_val = f"{seat_row}{seat_char} {name}" if name else f"{seat_row}{seat_char}"
                cell = ws2.cell(row=data_row, column=col, value=cell_val)
                cell.alignment = center_align
                cell.border = thin_border

                # 空座样式
                if not name:
                    set_empty_seat_style(cell)
                else:
                    cell.font = normal_font

            ws2.row_dimensions[data_row].height = 24

        # 更新当前行位置
        current_row = data_start_row + row_count

        # 车厢间空一行分隔
        current_row += 1

    # 冻结第一行
    ws2.freeze_panes = 'A2'

    # 3. 保存文件
    print(f"\nStep 3: Saving to {output_path}...")
    wb.save(output_path)
    print("SUCCESS!")
    return True


if __name__ == '__main__':
    if len(sys.argv) >= 3:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
    else:
        input_file = "0605D2987.xlsx"
        output_file = "output_seat_table.xlsx"

    try:
        generate_seat_excel(input_file, output_file)
    except Exception as e:
        print(f"\nERROR: {e}")
        print("\nTroubleshooting:")
        print("1. Make sure the file exists and is a valid Excel file")
        print("2. For old .xls files, install xlrd: pip install xlrd==1.2.0")
        print("3. Try opening and re-saving the file in Excel as .xlsx format")
        print("4. Check if the file contains the required columns: 姓名, 车厢, 座号")
        sys.exit(1)
